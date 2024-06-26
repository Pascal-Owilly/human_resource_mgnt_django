from django.shortcuts import render,redirect, resolve_url,reverse, get_object_or_404
from django.urls import reverse_lazy
from django.contrib.auth import get_user_model
from .models  import Employee, Department,Kin, Attendance, Leave, Recruitment, User, Admin, Client
from django.contrib.auth.views import LoginView
from django.contrib.auth import logout
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import FormView, CreateView,View,DetailView,TemplateView,ListView,UpdateView,DeleteView
from .forms import SuperuserRegistrationForm,EmployeeRegistrationForm,LoginForm,KinForm,DepartmentForm,AttendanceForm, LeaveForm, RecruitmentForm, ClientForm
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q

from django.contrib.auth.tokens import default_token_generator
from django.core.mail import send_mail
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.html import strip_tags

from django.contrib.auth.views import PasswordResetView, PasswordResetConfirmView, PasswordResetCompleteView
from django.views.generic import TemplateView
from datetime import datetime
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth import authenticate, login
from datetime import datetime, timedelta

class UserListView(View):
    template_name = 'hrms/users/user_list.html'
    paginate_by = 8  # Number of users per page

    def get(self, request):
        users_list = User.objects.filter(is_archived=False).order_by('-id')
        paginator = Paginator(users_list, self.paginate_by)
        page_number = request.GET.get('page')

        try:
            users = paginator.page(page_number)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page
            users = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results
            users = paginator.page(paginator.num_pages)

        return render(request, self.template_name, {'users': users})

class UserDetailView(View):
    template_name = 'hrms/users/user_detail.html'

    def get(self, request, user_id):
        user = get_object_or_404(User, pk=user_id)
        return render(request, self.template_name, {'user': user})

    def post(self, request, user_id):  # Ensure the method signature includes user_id
        user = get_object_or_404(User, pk=user_id)
        clockin_privileges = request.POST.get('clockin_privileges')
        if clockin_privileges:
            user.clockin_privileges = clockin_privileges
            user.save()
        return redirect('hrms:user_detail', user_id=user_id)

from .forms import UserUpdateForm

# User UpdateView

class UserUpdateView(LoginRequiredMixin, UpdateView):
    model = User
    form_class = UserUpdateForm
    template_name = 'hrms/users/user_update.html'
    success_url = reverse_lazy('hrms:user_list')
    login_url = 'hrms:login'

class UserArchiveView(LoginRequiredMixin, View):
    login_url = 'hrms:login'
    template_name = 'hrms/users/user_archive_confirm.html'  # Template for confirmation
    
    def get(self, request, pk, *args, **kwargs):
        user = get_object_or_404(User, pk=pk)
        return render(request, self.template_name, {'user': user})
    
    def post(self, request, pk, *args, **kwargs):
        user = get_object_or_404(User, pk=pk)
        user.is_archived = True
        user.save()
        return redirect(reverse_lazy('hrms:user_list'))

class UserUnarchiveView(LoginRequiredMixin, View):
    login_url = 'hrms:login'
    template_name = 'hrms/users/user_unarchive_confirm.html'  # Template for confirmation
    
    def get(self, request, pk, *args, **kwargs):
        user = get_object_or_404(User, pk=pk)
        return render(request, self.template_name, {'user': user})
    
    def post(self, request, pk, *args, **kwargs):
        user = get_object_or_404(User, pk=pk)
        user.is_archived = False
        user.save()
        return redirect(reverse_lazy('hrms:user_list'))

class ArchivedUserListView(ListView):
    template_name = 'hrms/users/archived_user_list.html'
    paginate_by = 8  # Number of users per page
    context_object_name = 'users'
    
    def get_queryset(self):
        return User.objects.filter(is_archived=True).order_by('-id')

# User DeleteView
class UserDeleteView(LoginRequiredMixin, DeleteView):
    model = User
    template_name = 'hrms/users/confirm_delete.html'
    success_url = reverse_lazy('hrms:user_list')
    login_url = 'hrms:login'

class Index(TemplateView):

   template_name = 'hrms/home/home.html'

def unauthorized(request):
    return render(request, 'auth/unauthorized.html')

def employee_dashboard(request):
    if request.user.role != 'employee' and not request.user.is_superuser:
        return redirect('unauthorized')
    return render(request, 'hrms/employees/employee_dashboard.html')


def send_password_reset_email(uidb64, token, email):
    reset_url = f"{settings.BASE_URL}{reverse('hrms:password_reset_confirm', kwargs={'uidb64': uidb64, 'token': token})}"
    subject = 'Set Your Password'
    message = f'Please click the following link to set your password: {reset_url}'
    sender_email = settings.DEFAULT_FROM_EMAIL

    send_mail(subject, message, sender_email, [email])

class CustomPasswordResetView(PasswordResetView):
    email_template_name = 'auth/password_reset_email.html'
    subject_template_name = 'auth/password_reset_subject.txt'
    template_name = 'auth/password_reset_form.html'
    success_url = reverse_lazy('hrms:password_reset_done')
    html_email_template_name = 'auth/password_reset_email.html'

    @staticmethod
    def send_password_reset_email(uidb64, token, email, first_name, last_name, username):
        # Construct the reset password URL
        reset_url = f"{settings.PROTOCOL}://{settings.DOMAIN}/reset/{uidb64}/{token}/"

        # Construct the email message
        subject = 'Set Your Password'
        context = {
            'reset_url': reset_url,
            'first_name': first_name,
            'last_name': last_name,
            'username': username,
            'theme_color': '#fdeb3d',
            'secondary_color': '#773697',
        }
        html_message = render_to_string('auth/password_reset_email.html', context)
        sender_email = settings.EMAIL_HOST_USER

        # Send the email
        send_mail(subject, None, sender_email, [email], html_message=html_message)

    def send_mail(self, subject_template_name, email_template_name, context, from_email, to_email, html_email_template_name=None):
        subject = render_to_string(subject_template_name, context)
        subject = ''.join(subject.splitlines())
        body = render_to_string(email_template_name, context)
        html_body = render_to_string(html_email_template_name, context)

        send_mail(subject, body, from_email, [to_email], html_message=html_body)

    def form_valid(self, form):
        email = form.cleaned_data['email']
        user = get_user_model().objects.get(email=email)
        uidb64 = urlsafe_base64_encode(force_bytes(user.pk))  
        token = default_token_generator.make_token(user)
        context = {
            'email': email,
            'domain': settings.DOMAIN,
            'site_name': settings.SITE_NAME,
            'uidb64': uidb64,
            'user': user,
            'token': token,
            'protocol': 'https' if self.request.is_secure() else 'http',
        }
        self.send_mail(
            self.subject_template_name,
            self.email_template_name,
            context,
            settings.DEFAULT_FROM_EMAIL,
            email,
            self.html_email_template_name,
        )
        
        return super().form_valid(form)

class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    success_url = reverse_lazy('hrms:password_reset_complete')
    template_name = 'auth/password_reset_confirm.html'

class CustomPasswordResetCompleteView(PasswordResetCompleteView):
    template_name = 'auth/password_reset_complete.html'
    success_url = reverse_lazy('hrms:password_reset_complete')

class CustomPasswordResetDoneView(TemplateView):
    template_name = 'auth/password_reset_done.html'

#   Authentication
class Register(CreateView):

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return render(request, 'auth/unauthorized.html')
        return super().dispatch(request, *args, **kwargs)

    model = get_user_model()
    form_class = SuperuserRegistrationForm
    template_name = 'hrms/registrations/register.html'
    success_url = reverse_lazy('hrms:login')

    def form_valid(self, form):
        user = form.save(commit=False)
        user.role = User.SUPERUSER
        user.is_superuser = True
        user.is_staff = True
        user.save()
        Admin.objects.create(admin=user)

        return redirect(self.success_url)

# def register_employee(request):
#     if request.method == 'POST':
#         form = EmployeeRegistrationForm(request.POST)
#         if form.is_valid():
#             user = form.save(commit=False)
#             user.role = User.Employee
#             user.save()

#             # Get the department from the form
#             department = form.cleaned_data.get('department')

#             # Generate uidb64 and token for password reset email
#             uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
#             token = default_token_generator.make_token(user)

#             # Send password reset email
#             send_password_reset_email(uidb64, token, user.email)

#             # Create a new Employee instance and associate the user with it
#             employee = Employee.objects.create(employee=user, department=department)
#             return redirect('employee_all')
#     else:
#         form = EmployeeRegistrationForm()
#     return render(request, 'hrms/employee/create.html', {'form': form})

class CustomLoginView(LoginView):
    template_name = 'hrms/registrations/login.html'
    authentication_form = LoginForm

    def form_valid(self, form):
        # This method is called when valid form data has been POSTed
        username = form.cleaned_data.get('username')
        password = form.cleaned_data.get('password')
        user = authenticate(self.request, username=username, password=password)

        if user is not None:
            login(self.request, user)
            # Determine user's role and redirect accordingly
            if user.role == User.SUPERUSER:
                return redirect('/dashboard/admin/')  # Redirect admin users to admin page
            elif user.role == User.EMPLOYEE:
                return redirect('/dashboard/attendance/emp/')  # Redirect employees to employee dashboard
            else:
                # Handle other roles or scenarios
                return redirect('/')  # Redirect to a generic dashboard
        else:
            # Add a non-field error to the form indicating invalid username or password
            form.add_error(None, 'Invalid username or password.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        # This method is called when the form is invalid
        return render(self.request, self.template_name, {'form': form})

class Login_View(LoginView):
    model = get_user_model()
    form_class = LoginForm
    template_name = 'hrms/registrations/login.html'

    def get_success_url(self):
        user = self.request.user
        url = '/'

        if user.role == User.SUPERUSER:
            url = reverse_lazy('hrms:admin_dashboard')
        elif user.role == User.ACCOUNT_MANAGER:
            url = reverse_lazy('hrms:account_manager_dashboard')
        elif user.role == User.EMPLOYEE:
            if User.ACCOUNT_MANAGER in user.roles.split(','):
                url = reverse_lazy('hrms:account_manager_dashboard')
            else:
                url = reverse_lazy('hrms:employee_dashboard')

        return url
class Logout_View(View):

    def get(self,request):
        logout(self.request)
        return redirect ('hrms:login',permanent=True)


# def export_management_dashboard(request):
#     if request.user.role != 'seller' and not request.user.is_superuser:
#         return redirect('unauthorized')
#     return render(request, 'export_management.html')

class AdminDashboard(LoginRequiredMixin, ListView):
    login_url = 'hrms:login'
    model = get_user_model()
    template_name = 'hrms/dashboard/index.html'
    context_object_name = 'qset'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return render(request, 'auth/unauthorized.html')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['emp_total'] = Employee.objects.all().count()
        context['dept_total'] = Department.objects.all().count()
        context['client_total'] = Client.objects.all().count()
        context['users_count'] = get_user_model().objects.all().count()
        context['admin_count'] = Admin.objects.all().count()
        context['workers'] = Employee.objects.filter(employee__is_archived=False).order_by('-id')
        return context

class AdminListView(View):
    template_name = 'admin/admin_list.html'
    paginate_by = 8  # Number of users per page

    def get(self, request):
        admin_list = Admin.objects.all().order_by('-id')
        paginator = Paginator(admin_list, self.paginate_by)
        page_number = request.GET.get('page')

        try:
            admins = paginator.page(page_number)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page
            admins = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results
            admins = paginator.page(paginator.num_pages)

        return render(request, self.template_name, {'admins': admins})

class Admin_View(LoginRequiredMixin,DetailView):
    queryset = Admin.objects.select_related('admin__department').order_by('-id')
    template_name = 'hrms/admin/single.html'
    context_object_name = 'amin'
    login_url = 'hrms:login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            query = Kin.objects.get(admin=self.object.pk)
            context["kin"] = query
            return context
        except ObjectDoesNotExist:
            return context

from .forms import AccountManagerRegistrationForm
from .models import AccountManager

class AccountManager_New(LoginRequiredMixin, CreateView):
    model = AccountManager
    form_class = AccountManagerRegistrationForm
    template_name = 'hrms/employee/create.html'
    login_url = 'hrms:login'
    redirect_field_name = 'redirect:'
    success_url = reverse_lazy('hrms:employee_all')  # URL to redirect to after a successful registration

    @staticmethod
    def send_password_reset_email(uidb64, token, email, first_name, last_name, username):
        # Construct the reset password URL
        reset_url = f"{settings.PROTOCOL}://{settings.DOMAIN}/reset/{uidb64}/{token}/"

        # Construct the email message
        subject = 'Set Your Password'
        context = {
            'reset_url': reset_url,
            'first_name': first_name,
            'last_name': last_name,
            'username': username,
            'theme_color': '#fdeb3d',
            'secondary_color': '#773697',
        }
        html_message = render_to_string('auth/password_reset_email.html', context)
        sender_email = settings.EMAIL_HOST_USER

        # Send the email
        send_mail(subject, None, sender_email, [email], html_message=html_message)

    def form_valid(self, form):

        # Generate a random password
        password = generate_random_password()

        # Save the form data
        user = form.save(commit=False)
        user.role = User.EMPLOYEE
        user.set_password(password)  # Set the random password
        user.save()

        # Get the department from the form
        department = form.cleaned_data.get('department')

        # Generate uidb64 and token for password reset email
        uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
        token = default_token_generator.make_token(user)

        # Get user details
        first_name = form.cleaned_data.get('first_name')
        last_name = form.cleaned_data.get('last_name')
        username = form.cleaned_data.get('username')

        # Send password reset email
        self.send_password_reset_email(uidb64, token, user.email, first_name, last_name, username)

        # Create a new Employee instance and associate the user with it
        AccountManager.objects.create(account_manager=user)

        return redirect(self.success_url)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.get_form()
        return context

class EmployeeDashboard(LoginRequiredMixin, ListView):
    login_url = 'hrms:login'
    model = get_user_model()
    template_name = 'hrms/employee/employee_dashboard.html'
    context_object_name = 'qset'

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'employee' and not request.user.is_superuser:
            return render(request, 'auth/unauthorized.html')
        return super().dispatch(request, *args, **kwargs)

# Employee's

import string
import random

def generate_random_password(length=12):
    characters = string.ascii_letters + string.digits + string.punctuation
    return ''.join(random.choice(characters) for _ in range(length))


class Employee_New(LoginRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeRegistrationForm
    template_name = 'hrms/employee/create.html'
    login_url = 'hrms:login'
    redirect_field_name = 'redirect:'
    success_url = reverse_lazy('hrms:employee_all')  # URL to redirect to after a successful registration

    @staticmethod
    def send_password_reset_email(uidb64, token, email, first_name, last_name, username):
        # Construct the reset password URL
        reset_url = f"{settings.PROTOCOL}://{settings.DOMAIN}/reset/{uidb64}/{token}/"

        # Construct the email message
        subject = 'Set Your Password'
        context = {
            'reset_url': reset_url,
            'first_name': first_name,
            'last_name': last_name,
            'username': username,
            'theme_color': '#fdeb3d',
            'secondary_color': '#773697',
        }
        html_message = render_to_string('auth/password_reset_email.html', context)
        sender_email = settings.EMAIL_HOST_USER

        # Send the email
        send_mail(subject, None, sender_email, [email], html_message=html_message)

    def form_valid(self, form):

        # Generate a random password
        password = generate_random_password()

        # Save the form data
        user = form.save(commit=False)
        user.role = User.EMPLOYEE
        user.set_password(password)  # Set the random password
        user.save()

        # Get the department from the form
        department = form.cleaned_data.get('department')

        # Generate uidb64 and token for password reset email
        uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
        token = default_token_generator.make_token(user)

        # Get user details
        first_name = form.cleaned_data.get('first_name')
        last_name = form.cleaned_data.get('last_name')
        username = form.cleaned_data.get('username')

        # Send password reset email
        self.send_password_reset_email(uidb64, token, user.email, first_name, last_name, username)

        # Create a new Employee instance and associate the user with it
        Employee.objects.create(employee=user)

        return redirect(self.success_url)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.get_form()
        return context

# Bulk reistration
import pandas as pd
from django.contrib.auth import get_user_model
from django.utils.crypto import get_random_string

# Define a function to perform bulk registration
from django.shortcuts import render
from .forms import UploadFileForm
from .utils import process_uploaded_file

def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            process_uploaded_file(file)
            return render(request, 'auth/employee_bulk_upload_success.html')
    else:
        form = UploadFileForm()
    return render(request, 'auth/employee_bulk_upload.html', {'form': form})


class Employee_All(LoginRequiredMixin, ListView):
    template_name = 'hrms/employee/index.html'
    model = Employee
    context_object_name = 'employees'
    paginate_by = 5
    login_url = 'hrms:login'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return render(request, 'auth/unauthorized.html')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(employee__is_archived=False).order_by('-id')

class Employee_View(LoginRequiredMixin,DetailView):
    queryset = Employee.objects.select_related('employee__department').order_by('-id')
    template_name = 'hrms/employee/single.html'
    context_object_name = 'employee'
    login_url = 'hrms:login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            query = Kin.objects.get(employee=self.object.pk)
            context["kin"] = query
            return context
        except ObjectDoesNotExist:
            return context


class Employee_Update(LoginRequiredMixin, View):
    template_name = 'hrms/employee/edit.html'
    login_url = 'hrms:login'

    def get(self, request, pk, *args, **kwargs):
        employee = Employee.objects.get(pk=pk)
        form = EmployeeRegistrationForm(instance=employee)
        return render(request, self.template_name, {'form': form})

    def post(self, request, pk, *args, **kwargs):
        employee = Employee.objects.get(pk=pk)
        form = EmployeeRegistrationForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('success_url')  # Replace 'success_url' with your desired redirect URL
        return render(request, self.template_name, {'form': form})

class Employee_Delete(LoginRequiredMixin, DeleteView):
    model = Employee
    template_name = 'hrms/employee/confirm_delete.html'
    success_url = reverse_lazy('hrms:employee_all')  # Replace 'success_url' with your desired redirect URL
    login_url = 'hrms:login'

class Employee_Kin_Add (LoginRequiredMixin,CreateView):
    model = Kin
    form_class = KinForm
    template_name = 'hrms/employee/kin_add.html'
    login_url = 'hrms:login'


    def get_context_data(self):
        context = super().get_context_data()
        if 'id' in self.kwargs:
            emp = Employee.objects.get(pk=self.kwargs['id'])
            context['emp'] = emp
            return context
        else:
            return context

class Employee_Kin_Update(LoginRequiredMixin,UpdateView):
    model = Kin
    form_class = KinForm
    template_name = 'hrms/employee/kin_update.html'
    login_url = 'hrms:login'

    def get_initial(self):
        initial = super(Employee_Kin_Update,self).get_initial()

        if 'id' in self.kwargs:
            emp =  Employee.objects.get(pk=self.kwargs['id'])
            initial['employee'] = emp.pk

            return initial

# Department views
class DepartmentListView(View):
    template_name = 'hrms/department/department_list.html'
    paginate_by = 5  # Number of users per page

    def get(self, request):
        department_list = Department.objects.all().order_by('-id')
        paginator = Paginator(department_list, self.paginate_by)
        page_number = request.GET.get('page')

        try:
            departments = paginator.page(page_number)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page
            departments = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results
            departments = paginator.page(paginator.num_pages)

        return render(request, self.template_name, {'departments': departments})

class Department_Detail(LoginRequiredMixin, ListView):
    context_object_name = 'employees'
    template_name = 'hrms/department/single.html'
    login_url = 'hrms:login'

    def get_queryset(self):
        department_pk = self.kwargs.get('pk')
        queryset = Employee.objects.filter(employee__department_id=department_pk)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        department_pk = self.kwargs.get('pk')
        department = get_object_or_404(Department, pk=department_pk)
        context["dept"] = department
        return context

class Department_New (LoginRequiredMixin,CreateView):
    model = Department
    template_name = 'hrms/department/create.html'
    form_class = DepartmentForm
    login_url = 'hrms:login'

class Department_Update(LoginRequiredMixin,UpdateView):
    model = Department
    template_name = 'hrms/department/edit.html'
    form_class = DepartmentForm
    login_url = 'hrms:login'
    success_url = reverse_lazy('hrms:dashboard')

# Client views 

class ClientListView(View):
    template_name = 'hrms/client/client_list.html'
    paginate_by = 5  # Number of users per page

    def get(self, request):
        client_list = Client.objects.all().order_by('-id')
        paginator = Paginator(client_list, self.paginate_by)
        page_number = request.GET.get('page')

        try:
            clients = paginator.page(page_number)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page
            clients = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results
            clients = paginator.page(paginator.num_pages)

        return render(request, self.template_name, {'clients': clients})

class Client_Detail(LoginRequiredMixin, ListView):
    context_object_name = 'clients'
    template_name = 'hrms/client/single.html'
    login_url = 'hrms:login'

    def get_queryset(self):
        department_pk = self.kwargs.get('pk')
        queryset = Client.objects.filter(department_id=department_pk)
        return queryset


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        client_pk = self.kwargs.get('pk')
        client = get_object_or_404(Client, pk=client_pk)
        context["clnt"] = Client
        return context

class Client_New(LoginRequiredMixin, CreateView):
    model = Client
    template_name = 'hrms/client/create.html'
    form_class = ClientForm
    login_url = 'hrms:login'

    def form_valid(self, form):
        # Add debugging output
        print("Form is valid. Cleaned data:", form.cleaned_data)
        return super().form_valid(form)

    def form_invalid(self, form):
        # Add debugging output
        print("Form is invalid. Errors:", form.errors)
        return super().form_invalid(form)

class Client_Update(LoginRequiredMixin,UpdateView):
    model = Client
    template_name = 'hrms/client/edit.html'
    form_class = ClientForm
    login_url = 'hrms:login'
    success_url = reverse_lazy('hrms:dashboard')

    def get_object(self, queryset=None):
        # handle case where client does not exist
        return get_object_or_404(Client, pk=self.kwargs.get('pk'))


#Attendance View

from django.utils import timezone
from datetime import datetime

from django.core.paginator import Paginator

class Attendance_Admin(LoginRequiredMixin, View):
    login_url = 'hrms:login'

    def dispatch(self, request, *args, **kwargs):
            if not request.user.is_superuser:
                return render(request, 'auth/unauthorized.html')
            return super().dispatch(request, *args, **kwargs)


    def get(self, request, *args, **kwargs):
        # Get query parameters
        date = request.GET.get('date')
        keyword = request.GET.get('keyword')
        geofence_center = (-1.315638, 36.862129)  # Example: Nairobi coordinates

        # Retrieve attendance records based on the provided date
        if date:
            try:
                selected_date = datetime.strptime(date, '%Y-%m-%d').date()
                present_staffers = Attendance.objects.filter(Q(status='PRESENT') & Q(date=selected_date)).order_by('-id')
            except ValueError:
                selected_date = None
                present_staffers = Attendance.objects.none()
        else:
            selected_date = timezone.localdate()
            present_staffers = Attendance.objects.filter(Q(status='PRESENT') & Q(date=selected_date)).order_by('-id')

        # Perform search if keyword is provided
        if keyword:
            present_staffers = present_staffers.filter(
                Q(staff__admin__first_name__icontains=keyword) |
                Q(staff__admin__last_name__icontains=keyword)
            )

        # Calculate distance for each present staffer
        for staff in present_staffers:
            if staff.latitude and staff.longitude:
                staff.distance = geodesic((staff.latitude, staff.longitude), geofence_center).meters
            else:
                staff.distance = None

        # Pagination
        paginator = Paginator(present_staffers, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        admin = Admin.objects.get(admin=request.user)

        clocked_in = Attendance.objects.filter(admin=admin, date=timezone.localdate(), last_out__isnull=True).exists()

        context = {
            'today': timezone.localdate(),
            'present_staffers': page_obj,
            'selected_date': selected_date,
            'keyword': keyword,
            'page_obj': page_obj,
            'clocked_in': clocked_in
        }

        return render(request, 'hrms/attendance/create.html', context)


class Attendance_Employee(LoginRequiredMixin, View):
    login_url = 'hrms:login'

    def get(self, request, *args, **kwargs):
        # Get query parameters
        date_emp = request.GET.get('date')
        keyword = request.GET.get('keyword')
        geofence_center = (-1.315638, 36.862129)  # Example: Nairobi coordinates

        # Retrieve the logged-in employee
        employee = Employee.objects.get(employee=request.user)

        # Determine the start and end of the current week (Monday to Friday)
        today_emp = timezone.localdate()
        start_of_week = today_emp - timedelta(days=today_emp.weekday())  # Monday
        end_of_week = start_of_week + timedelta(days=4)  # Friday

        # Retrieve attendance records for the logged-in employee within the current week
        if date_emp:
            try:
                selected_date = datetime.strptime(date_emp, '%Y-%m-%d').date()
                present_employees = Attendance.objects.filter(
                    Q(status='PRESENT') & Q(date=selected_date) & Q(staff=employee)
                ).order_by('-id')
            except ValueError:
                selected_date = None
                present_employees = Attendance.objects.none()
        else:
            selected_date = None
            present_employees = Attendance.objects.filter(
                Q(status='PRESENT') & Q(date__range=(start_of_week, end_of_week)) & Q(staff=employee)
            ).order_by('-id')

        # Perform search if keyword is provided
        if keyword:
            present_employees = present_employees.filter(
                Q(staff__employee__first_name__icontains=keyword) |
                Q(staff__employee__last_name__icontains=keyword)
            )

        # Calculate distance for each present staffer
        for staff in present_employees:
            if staff.latitude and staff.longitude:
                staff.distance = geodesic((staff.latitude, staff.longitude), geofence_center).meters
            else:
                staff.distance = None

        # Pagination
        paginator = Paginator(present_employees, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Check if the logged-in employee is clocked in
        clocked_in = Attendance.objects.filter(
            staff=employee, date=timezone.localdate(), last_out__isnull=True
        ).exists()

        context = {
            'today_emp': today_emp,
            'present_employees': page_obj,
            'selected_date': selected_date,
            'keyword': keyword,
            'page_obj': page_obj,
            'clocked_in': clocked_in
        }

        return render(request, 'hrms/employee/employee_dashboard.html', context)


from geopy.distance import geodesic

class Attendance_Out(LoginRequiredMixin, View):
    login_url = 'hrms:login'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return render(request, 'auth/unauthorized.html')
        return super().dispatch(request, *args, **kwargs)


    def get(self, request, *args, **kwargs):
        try:
            user = Attendance.objects.get(
                Q(staff__id=self.kwargs['user_id']) &
                Q(status='PRESENT') &
                Q(date=timezone.localdate())
            )
            user.last_out = timezone.localtime()
            user.save()
            return redirect('hrms:attendance_new')
        except Attendance.DoesNotExist:
            return redirect('hrms:attendance_new')

class Attendance_Out_Emp(LoginRequiredMixin, View):
    login_url = 'hrms:login'

    def get(self, request, *args, **kwargs):
        try:
            user = Attendance.objects.get(
                Q(staff__id=self.kwargs['user_id']) &
                Q(status='PRESENT') &
                Q(date=timezone.localdate())
            )
            user.last_out = timezone.localtime()
            user.save()
            return redirect('hrms:attendance_employee')
        except Attendance.DoesNotExist:
            return redirect('hrms:attendance_employee')

from django.contrib import messages
from django.http import JsonResponse
from django.utils import timezone
from geopy.distance import geodesic
from .models import Attendance, Employee  # Ensure your models are imported

class AdminClockInView(LoginRequiredMixin, View):
    login_url = 'hrms:login'

    def post(self, request, *args, **kwargs):
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')
        print(f"Received coordinates: latitude={latitude}, longitude={longitude}")

        if latitude is None or longitude is None:
            messages.error(request, 'Latitude and Longitude are required.')
            return redirect('hrms:attendance_employee')

        # Retrieve the logged-in user's employee instance
        try:
            admin = Admin.objects.get(admin=request.user)
        except Admin.DoesNotExist:
            messages.error(request, 'You are not an admin')
            return redirect('hrms:attendance_new')

        # Define the geofence center and radius
        # geofence_center = (-1.2540381172761967, 36.71374983009918)  # Example: main institution coordinates
        geofence_center = (-1.2504447, 36.7150981)
        # main office onesmus phone -1.33220199, 36.8622648

        # marials phone -1.2974781588016533, 36.76480694390815
        # my phone coordinates -1.2504447, 36.7150981
        # marials coordinates -1.2975869, 36.7649746
        # -1.2558260124069696, 36.69340338379849 Kinoo jacmin
        # geofence_center = (-1.2829549455322522, 36.82593840499116)  # Example: Nairobi odeon coordinates
        # Naivas uthiru coperation (-1.2606487526706478, 36.709970821254515)
        # Retrieve attendance records based on the provided date
        # Uthiru chiefs camp -1.2540381172761967, 36.71374983009918
        # Kisumu international airport -0.08182281166829138, 34.72939625715378
        # kabete national polytechnic -1.263690553734057, 36.72265660553466
        # Mombasa -4.0454093873302055, 39.65720790666958

        # Kangemi -1.2710588266841318, 36.739521489451775
        # kinoo hse -1.2841, 36.8155
        # ABC PLACE -1.2584232411933278, 36.77113775720258
        # Westlands -1.2676956311569731, 36.81221729528121
        # Riara road -1.297610, 36.764904

        geofence_radius_km = 0.3  # 200 meters (0.2 km)

        try:
            admin_location = (float(latitude), float(longitude))
            distance_km = geodesic(admin_location, geofence_center).km
            print(f"Calculated distance: {distance_km} km")
        except ValueError:
            messages.error(request, 'Invalid latitude or longitude. Please ensure that your location services are enabled')
            return redirect('hrms:attendance_new')

        # Check if the user has the privilege to clock in from anywhere
        if request.user.clockin_privileges == User.CAN_CLOCK_IN_ANYWHERE:
            self.clock_in(admin, latitude, longitude, distance_km, request)
            return redirect('hrms:attendance_new')
        else:
            # Check if the employee is within the geofence area
            if distance_km <= geofence_radius_km:
                self.clock_in(admin, latitude, longitude, distance_km, request)
            else:
                messages.error(request, f'You are outside the allowed geofence area. Latitude: {latitude}, Longitude: {longitude}. Distance from geofence center: {distance_km:.2f} km')

        return redirect('hrms:attendance_new')

    def clock_in(self, admin, latitude, longitude, distance_km, request):
        # Check if the employee is already clocked in
        attendance = Attendance.objects.filter(admin=admin, date=timezone.localdate(), last_out__isnull=True).first()
        if attendance:
            # Clocking out
            attendance.last_out = timezone.localtime()
            attendance.save()
            messages.success(request, f'Clock-out successful! Latitude: {latitude}, Longitude: {longitude}. Distance from geofence center: {distance_km:.2f} km')

        else:
            # Clocking in
            Attendance.objects.create(
                admin=admin,
                latitude=latitude,
                longitude=longitude,
                first_in=timezone.localtime(),
                status='PRESENT'
            )
            self.send_late_arrival_notification(admin, request)
            messages.success(request, f'Clock-in successful! Latitude: {latitude}, Longitude: {longitude}. Distance from geofence center: {distance_km:.2f} km')

    def send_late_arrival_notification(self, admin, request):
        attendance_time = timezone.localtime()
        if attendance_time.time() > datetime.strptime('08:30', '%H:%M').time():
            subject = 'Late Clock-in Notification'
            html_message = render_to_string('hrms/employee/employee_late_arrival.html', {
                'first_name': admin.admin.first_name,
                'last_name': admin.admin.last_name,
                'username': admin.admin.username,
                'clock_in_time': attendance_time.strftime("%H:%M:%S")
            })
            plain_message = strip_tags(html_message)
            from_email = settings.DEFAULT_FROM_EMAIL
            # to_email = 'bollo.j@jawabubest.co.ke'
            to_email = 'pascalouma55@gmail.com'

            send_mail(
                subject,
                plain_message,
                from_email,
                [to_email],
                html_message=html_message,
                fail_silently=False,
            )

class EmployeeClockInView(LoginRequiredMixin, View):
    login_url = 'hrms:login'

    def post(self, request, *args, **kwargs):
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')
        print(f"Received coordinates: latitude={latitude}, longitude={longitude}")

        if latitude is None or longitude is None:
            messages.error(request, 'Latitude and Longitude are required.')
            return redirect('hrms:attendance_employee')

        # Retrieve the logged-in user's employee instance
        try:
            employee = Employee.objects.get(employee=request.user)
        except Employee.DoesNotExist:
            messages.error(request, 'You are not an employee')
            return redirect('hrms:attendance_employee')

        # Define the geofence center and radius
        # geofence_center = (-1.315638, 36.862129)  # Example: main institution coordinates
        geofence_center = (-1.2504447, 36.7150981)
        # james phone -1.3319954, 36.8622605
        # marials phone -1.2974781588016533, 36.76480694390815
        # my phones -1.2514447, 36.7150981
        # marialsplace -1.2975869, 36.7649746   
        # my place -1.2841, 36.8155
        # -1.2558260124069696, 36.69340338379849 Kinoo jacmin
        # geofence_center = (-1.2829549455322522, 36.82593840499116)  # Example: Nairobi odeon coordinates
        # Naivas uthiru coperation (-1.2606487526706478, 36.709970821254515)
        # Retrieve attendance records based on the provided date
        # Uthiru chiefs camp -1.2540381172761967, 36.71374983009918
        # Kisumu international airport -0.08182281166829138, 34.72939625715378
        # kabete national polytechnic -1.263690553734057, 36.72265660553466
        # Mombasa -4.0454093873302055, 39.65720790666958

        # Kangemi -1.2710588266841318, 36.739521489451775
        # kinoo hse -1.2841, 36.8155
        # ABC PLACE -1.2584232411933278, 36.77113775720258
        # Westlands -1.2676956311569731, 36.81221729528121

        geofence_radius_km = 0.1  # 200 meters (0.2 km)

        try:
            employee_location = (float(latitude), float(longitude))
            distance_km = geodesic(employee_location, geofence_center).km
            print(f"Calculated distance: {distance_km} km")
        except ValueError:
            messages.error(request, 'Invalid latitude or longitude. Please ensure that your location services are enabled')
            return redirect('hrms:attendance_employee')

        # Check if the user has the privilege to clock in from anywhere
        if request.user.clockin_privileges == User.CAN_CLOCK_IN_ANYWHERE:
            self.clock_in(employee, latitude, longitude, distance_km, request)
            return redirect('hrms:attendance_employee')
        else:
            # Check if the employee is within the geofence area
            if distance_km <= geofence_radius_km:
                self.clock_in(employee, latitude, longitude, distance_km, request)
            else:
                messages.error(request, f'You are outside the allowed geofence area. Latitude: {latitude}, Longitude: {longitude}. Distance from geofence center: {distance_km:.2f} km')

        return redirect('hrms:attendance_employee')

    def clock_in(self, employee, latitude, longitude, distance_km, request):
        # Check if the employee is already clocked in
        attendance = Attendance.objects.filter(staff=employee, date=timezone.localdate(), last_out__isnull=True).first()
        if attendance:
            # Clocking out
            attendance.last_out = timezone.localtime()
            attendance.save()
            messages.success(request, f'Clock-out successful! Latitude: {latitude}, Longitude: {longitude}. Distance from geofence center: {distance_km:.2f} km')

        else:
            # Clocking in
            Attendance.objects.create(
                staff=employee,
                latitude=latitude,
                longitude=longitude,
                first_in=timezone.localtime(),
                status='PRESENT'
            )
            self.send_late_arrival_notification(employee, request)
            messages.success(request, f'Clock-in successful! Latitude: {latitude}, Longitude: {longitude}. Distance from geofence center: {distance_km:.2f} km')


    def send_late_arrival_notification(self, employee, request):
        attendance_time = timezone.localtime()
        if attendance_time.time() > datetime.strptime('08:30', '%H:%M').time():
            subject = 'Late Clock-in Notification'
            html_message = render_to_string('hrms/employee/employee_late_arrival.html', {
                'first_name': employee.employee.first_name,
                'last_name': employee.employee.last_name,
                'username': employee.employee.username,
                'clock_in_time': attendance_time.strftime("%H:%M:%S")
            })
            plain_message = strip_tags(html_message)
            from_email = settings.DEFAULT_FROM_EMAIL
            # to_email = 'bollo.j@jawabubest.co.ke'
            to_email = 'pascalouma55@gmail.com'

            send_mail(
                subject,
                plain_message,
                from_email,
                [to_email],
                html_message=html_message,
                fail_silently=False,
            )


from geopy.distance import geodesic

from django.contrib import messages
from django.http import JsonResponse
from django.utils import timezone
from geopy.distance import geodesic
from .models import Attendance, Employee

import io
from django.http import HttpResponse
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
import openpyxl
from .models import Attendance
from django.utils import timezone

class DownloadPDF(View):
    def get(self, request, *args, **kwargs):
        date = request.GET.get('date', timezone.localdate())
        keyword = request.GET.get('keyword', '')

        attendances = Attendance.objects.filter(
            date=date
        )

        if keyword:
            attendances = attendances.filter(
                Q(staff__employee__first_name__icontains=keyword) |
                Q(staff__employee__last_name__icontains=keyword)
            )

        template = get_template('hrms/attendance/download_data/pdf_template.html')
        context = {
            'attendances': attendances,
            'date': date,
            'keyword': keyword,
        }
        html = template.render(context)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="attendance.pdf"'

        pisa_status = pisa.CreatePDF(
            io.BytesIO(html.encode('UTF-8')),
            dest=response,
        )

        if pisa_status.err:
            return HttpResponse('We had some errors with your request', status=500)
        return response

class DownloadExcel(View):
    def get(self, request, *args, **kwargs):
        date = request.GET.get('date', timezone.localdate())
        keyword = request.GET.get('keyword', '')

        attendances = Attendance.objects.filter(
            date=date
        )

        if keyword:
            attendances = attendances.filter(
                Q(staff__employee__first_name__icontains=keyword) |
                Q(staff__employee__last_name__icontains=keyword)
            )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="attendance.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Attendance'

        columns = ['Date', 'First-In (Arrival)', 'Last-Out (Departure)', 'Name', 'Distance (m)']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for attendance in attendances:
            row_num += 1
            row = [
                attendance.date,
                attendance.first_in,
                attendance.last_out,
                f"{attendance.staff.employee.first_name} {attendance.staff.employee.last_name}",
                # f"{attendance.distance if attendance.distance else 'N/A'} m",
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response


class LeaveNew (LoginRequiredMixin,CreateView, ListView):
    model = Leave
    template_name = 'hrms/leave/create.html'
    form_class = LeaveForm
    login_url = 'hrms:login'
    success_url = reverse_lazy('hrms:leave_new')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["leaves"] = Leave.objects.all()
        return context

class Payroll(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'hrms/payroll/index.html'
    login_url = 'hrms:login'
    context_object_name = 'stfpay'

class RecruitmentNew (CreateView):
    model = Recruitment
    template_name = 'hrms/recruitment/index.html'
    form_class = RecruitmentForm
    success_url = reverse_lazy('hrms:recruitment')

class RecruitmentAll(LoginRequiredMixin,ListView):
    model = Recruitment
    login_url = 'hrms:login'
    template_name = 'hrms/recruitment/all.html'
    context_object_name = 'recruit'

class RecruitmentDelete (LoginRequiredMixin,View):
    login_url = 'hrms:login'
    def get (self, request,pk):
     form_app = Recruitment.objects.get(pk=pk)
     form_app.delete()
     return redirect('hrms:recruitmentall', permanent=True)

class Pay(LoginRequiredMixin,ListView):
    model = Employee
    template_name = 'hrms/payroll/index.html'
    context_object_name = 'emps'
    login_url = 'hrms:login'
